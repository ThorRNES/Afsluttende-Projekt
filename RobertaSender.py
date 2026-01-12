from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, UTC
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


""" def json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value) """
# udkommenteret ovenstående og tilføjet nedenstående for at håndtere UTC tid korrekt

def json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime) and value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        return value.isoformat().replace("+00:00", "Z")
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def excel_to_records(xlsx_path: str, sheet_name: Optional[str] = None) -> Tuple[str, List[Dict[str, Any]]]:
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        actual_sheet = sheet_name
    else:
        ws = wb.active
        actual_sheet = ws.title

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return actual_sheet, []

    headers: List[str] = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(header_row, start=1):
        name = (str(h).strip() if h is not None and str(h).strip() else f"col_{i}")
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        headers.append(name)

    records: List[Dict[str, Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue
        rec: Dict[str, Any] = {}
        for idx, key in enumerate(headers):
            rec[key] = row[idx] if idx < len(row) else None
        records.append(rec)

    return actual_sheet, records


def post_json(endpoint_url: str, payload: Dict[str, Any], timeout_s: float = 20.0) -> Dict[str, Any]:
    data = json.dumps(payload, default=json_default).encode("utf-8")
    req = urllib.request.Request(
        endpoint_url,
        data=data,
        method="POST",
        headers={"Content-Type": "application/json", "User-Agent": "RobotA/1.0"},
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout_s) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            try:
                return json.loads(body) if body else {"ok": True}
            except json.JSONDecodeError:
                return {"ok": True, "raw_response": body}
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code}: {err_body}") from e
    except urllib.error.URLError as e:
        raise RuntimeError(f"Connection error: {e}") from e


def build_payload(excel_path: str, sheet: str, records: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "source": {
            "excel_path": os.path.abspath(excel_path),
            "excel_filename": os.path.basename(excel_path),
            "sheet": sheet,
        },
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "records": records,
    }


def main() -> int:
    DEFAULT_EXCEL = r"C:/Users/thorn/OneDrive/Skrivebord/Datamatiker Afsluttende projekt/Produkt/Test cases/Ark Database.xlsx" #Hardcoded sti til at finde Excel ark
    DEFAULT_ENDPOINT = "http://127.0.0.1:8080/ingest" 

    parser = argparse.ArgumentParser(description="Robot A: Excel -> JSON -> send to Robot B.")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Path to local .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    parser.add_argument("--endpoint", default=DEFAULT_ENDPOINT, help="Robot B ingest URL, e.g. http://127.0.0.1:8080/ingest")
    parser.add_argument("--timeout", type=float, default=20.0, help="HTTP timeout seconds")
    parser.add_argument("--save-json", default=None, help="Optional path to save JSON locally before sending")
    args = parser.parse_args()

    if not os.path.isfile(args.excel):
        print(f"Excel file not found: {args.excel}", file=sys.stderr)
        return 2

    try:
        actual_sheet, records = excel_to_records(args.excel, args.sheet)
        payload = build_payload(args.excel, actual_sheet, records)

        if args.save_json:
            with open(args.save_json, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, ensure_ascii=False, default=json_default)

        resp = post_json(args.endpoint, payload, timeout_s=args.timeout)
        print(json.dumps(resp, indent=2, ensure_ascii=False))
        return 0
    except Exception as e:
        print(f"Robot A failed: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())